import os
import datetime

from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active

if(os.path.isfile(os.path.abspath(os.getcwd())+"/Data.xlsx")):
    wb = load_workbook('Data.xlsx')
    ws = wb["Sensor1"]
else:
    wb = Workbook()
    ws = wb.create_sheet("Sensor1")
    OptionList = ["Sensor 1"]
    ws.append(['Temperature', 'Humidity', 'Soil Moisture', 'Light Level', 'Date Time'])
    ws.append([0, 0, 0, 0,datetime.datetime.now()])
    ws2 = wb["Sheet"]
    ws2.append(['deviceID', 'AccessToken', 'DeviceName', 'Soil Moisture Threshold'])
    ws2.append(['<Your Device ID>', '<Your Device Access token>', 'Sensor1', '200'])
    wb.save('Data.xlsx')
