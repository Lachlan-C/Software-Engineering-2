import requests
import os
import datetime
import time

from openpyxl import Workbook
from openpyxl import load_workbook

def GetData():
    wb = load_workbook('Data.xlsx')
    ws1 = wb["Sheet"]
    for each in ws1.iter_rows(min_row=2):
        print(f"Geting Data for {each[2].value}")
        temp = requests.get(f"https://api.particle.io/v1/devices/{each[0].value}/temperature?access_token={each[1].value}")
        hum = requests.get(f"https://api.particle.io/v1/devices/{each[0].value}/humidity?access_token={each[1].value}")
        soil = requests.get(f"https://api.particle.io/v1/devices/{each[0].value}/SoilMoisture?access_token={each[1].value}")
        light = requests.get(f"https://api.particle.io/v1/devices/{each[0].value}/Light?access_token={each[1].value}")
        print(f"Data Collected for {each[2].value}")
        if (temp.json().get('result') != None and hum.json().get('result') != None and soil.json().get('result') != None and light.json().get('result') != None ):
            ws =  wb[each[2].value]
            ws.append([temp.json().get('result'), hum.json().get('result'), soil.json().get('result'), light.json().get('result'),datetime.datetime.now()])
            wb.save('Data.xlsx')
            print(f"Data Saved for {each[2].value}")
        else:
            print(f"Data Error for {each[2].value} data not saved")

while True:
    GetData()
    time.sleep(300)
