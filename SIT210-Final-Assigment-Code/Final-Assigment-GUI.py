from tkinter import *
from tkinter.ttk import *

import matplotlib
matplotlib.use("TkAgg")

from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import time

from functools import partial

import requests
import os
import datetime

from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('Data.xlsx')

def GetThreshold(wb, deviceName):
    ws1 = wb["Sheet"]
    for each in ws1.iter_rows(min_row=2):
        if (each[2].value == deviceName):
            return int(each[3].value)
        else:
            return 0
        
def SetThreshold(wb, deviceName, threshold):
    ThresholdEntered = threshold.get()
    ws1 = wb["Sheet"]
    for each in ws1.iter_rows(min_row=2):
        if (each[2].value == deviceName):
            r = requests.post(f"https://api.particle.io/v1/devices/{each[0].value}/MoistureThreshold", data = {'args': ThresholdEntered, 'access_token' : each[1].value})
            each[3].value = ThresholdEntered
            wb.save('Data.xlsx')
            
def GetXandY(ws, column):
    x = []
    y = []
    for row in ws.iter_rows(min_row=3): 
        y.append(row[column].value)
        x.append(row[4].value)
    return x, y

def Average(x): 
    return sum(x) / len(x) 
       
def AddNewSensor(wb,DeviceID, Token,name):
    wsnew = wb.create_sheet(name)
    wsnew.append(['Temperature', 'Humidity', 'Soil Moisture', 'Light Level', 'Date Time'])
    wsnew.append([0, 0, 0, 0,datetime.datetime.now()])
    ws2 = wb["Sheet"]
    ws2.append(['deviceID', 'AccessToken', 'DeviceName', 'Soil Moisture Threshold'])
    ws2.append([DeviceID, Token, name, '200'])
    wb.save('Data.xlsx')
    
def RemoveASensor(wb,name):
    wsdelete = wb[name]
    wb.remove(wsdelete)
    print("Sensor Data DELETED")
    ws2 = wb["Sheet"]
    for row in ws2.iter_rows(min_row=2):
        if (row[2].value == name):
            ws2.delete_rows(row[0].row,1)
            print("Device Removed")
    wb.save('Data.xlsx')


def menu():
    wb = load_workbook('Data.xlsx')
    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
    print("Options")
    print("1. View a Device")
    print("2. Add new Device")
    print("3. Remove a Device")
    print("4. Quit")
    choice = input("Enter your choice: ")
    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
    if (choice == "1"):
        listsensors = wb.sheetnames
        listsensors.pop(0)
        print(listsensors)
        device = input("Enter name of the device: ")
        for sheet in listsensors:
            if (sheet == device and device != "Sheet"):
                view(device)
        print("Device Not Found")
        menu()
    elif (choice == "2"):
        name = input("Enter name of the new device: ")
        id_input = input("Enter the Argons ID: ")
        Token = input("Enter the argons Access ID: ")
        AddNewSensor(wb, id_input, Token,name)
        print("New Device Added")
        menu()
    elif (choice == "3"):
        listsensors = wb.sheetnames
        listsensors.pop(0)
        print(listsensors)
        name = input("Enter name of the device to remove: ")
        for sheet in listsensors:
            if (sheet == name and name != "Sheet"):
                RemoveASensor(wb,name)
            else:
                print("Device Not Found")
        menu()
    elif (choice == "4"):
        pass
    else:
        print("Enter Valid Option")
        menu()
        
def view(selectedDevice):
    win = Tk()
    wb = load_workbook('Data.xlsx')
    
    ws1 = wb["Sheet"]

    win.title("Plant Monitoring System")
            
    ws = wb[selectedDevice]

    for each in ws1.iter_rows(min_row=2):
        if (each[2].value == selectedDevice):
            threshold = each[3].value

    labeldevice = Label(win, text=selectedDevice)
    labeldevice.grid(row=0,column=0)
            
    label = Label(win, text='Enter Moisture Threshold')
    label.grid(row=2,column=0)

    ThresholdEnter = StringVar()
    ThresholdEntry = Entry(win, width=10, textvariable=ThresholdEnter)
    ThresholdEntry.grid(row=3,column=0)
        
    SubmitThreshold = Button(win, text='Submit Threshold',command=partial(SetThreshold, wb, selectedDevice,ThresholdEnter), width=24)
    SubmitThreshold.grid(row=4,column=0)
            
    CurrentMoisture = ws.cell(row=ws.max_row, column=3).value
    thresh = GetThreshold(wb, selectedDevice)
    watervalue = Label(win, text=f'Current Moisture: {CurrentMoisture} \nMoisture Level needed: {thresh}')
    if (CurrentMoisture < thresh):
        Water = Label(win, text=f'{selectedDevice} needs Watering')
    else:
        Water = Label(win, text=f'{selectedDevice} is Hydrated')
    Water.grid(row=6, column=0)
    watervalue.grid(row=7, column=0)
            
    templabel = Label(win, text='Temperature vs Time')
    templabel.grid(row=0,column=2)
    tempfig = Figure(figsize=(8, 3))
    ax = tempfig.add_subplot(111)
    x,y = GetXandY(ws, 0)
    ax.plot(x, y)
    tempcanvas = FigureCanvasTkAgg(tempfig, win)
    tempcanvas.get_tk_widget().grid(row = 1, column = 2)
    tempAverage = Label(win, text=f'Average: {Average(y)}')
    tempAverage.grid(row=2,column=2)
            
    humlabel = Label(win, text='Humidity vs Time')
    humlabel.grid(row=4,column=2)
    humfig = Figure(figsize=(8, 3))
    bx = humfig.add_subplot(111)
    x,y = GetXandY(ws, 1)
    bx.plot(x,y)
    humcanvas = FigureCanvasTkAgg(humfig, win)
    humcanvas.get_tk_widget().grid(row = 5, column = 2)
    humAverage = Label(win, text=f'Average: {Average(y)}')
    humAverage.grid(row=6,column=2)
            
    soillabel = Label(win, text='Soil Moisture vs Time')
    soillabel.grid(row=0,column=3)
    soilfig = Figure(figsize=(8, 3))
    cx = soilfig.add_subplot(111)
    x,y = GetXandY(ws, 2)
    cx.plot(x,y)
    soilcanvas = FigureCanvasTkAgg(soilfig, win)
    soilcanvas.get_tk_widget().grid(row = 1, column = 3)
    soilAverage = Label(win, text=f'Average: {Average(y)}')
    soilAverage.grid(row=2,column=3)
            
    lightlabel = Label(win, text='Light Level vs Time')
    lightlabel.grid(row=4,column=3)
    lightfig = Figure(figsize=(8, 3))
    dx = lightfig.add_subplot(111)
    x,y = GetXandY(ws, 3)
    dx.plot(x,y)
    lightcanvas = FigureCanvasTkAgg(lightfig, win)
    lightcanvas.get_tk_widget().grid(row = 5, column = 3)
    lightAverage = Label(win, text=f'Average: {Average(y)}')
    lightAverage.grid(row=6,column=3)

    def closing():
        win.destroy()
        menu()

    win.protocol("WM_DELETE_WINDOW", closing)
    win.mainloop()
menu()